import pandas as pd
from openpyxl import load_workbook
class Person: 
    def __init__ (self, name):
        self.name = name
        self.points = 0

    def attended_Event(self):
        self.points+=5
    def participated_Event(self):
        self.points+=3
    def cleaned_Up(self):
        self.points+=1


def main(): 
    # Read the Excel file
    df = pd.read_excel("data.xlsx")


    # to get cell color
    wb = load_workbook("data.xlsx")
    ws = wb.active


    # checking if cell is green
    def did_rsvp(cell):
        return cell.fill.fgColor.rgb == "FF00FF00"



    # finding a cell location by the name in the cell
    def find_cell(val):
        for row in ws.iter_rows(min_row = 1, max_row = 91, min_col = 3, max_col = 3):
            for cell in row: 
                if cell.value == val: 
                    return cell
        return None


    # Extract the name column
    person_column_name = df.columns[2]  # Assuming the third column is at index 2 (0-based index)
    person_column_data = df[person_column_name]

    # Filter the non-empty cells from both columns
    non_empty_third_column = person_column_data.dropna().iloc[:91]

    # Holds who attended and was absent 
    attended = []
    absent = []


    # adding to arrays
    for cell in non_empty_third_column:
        spec_cel = find_cell(cell)
        if did_rsvp(spec_cel):
            person = Person(cell)
            person.attended_Event()
            attended.append(person)
        elif spec_cel == None:
            print("DNE")
        else:
            abPerson = Person(cell)
            absent.append(abPerson)
        

    # Printing the arrays

    print("People that RSVP'ed and Attended: ")
    for person in attended:
        print(f"{person.name}: {person.points}")
    print() 
    print("People that were absent")
    for person in absent: 
        print(f"{person.name}")


if __name__ == "__main__":
    main()
